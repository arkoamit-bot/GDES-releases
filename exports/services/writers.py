"""File writers for research dataset export.

Each writer accepts (headers, rows) as produced by ``dataset.build_dataset``
and returns bytes (or text for CSV).  These are thin wrappers around
third-party libraries openpyxl (Excel) and pyreadstat (SPSS).
"""
from __future__ import annotations

import csv
import io
from typing import Any


def to_csv(
    headers: list[str],
    rows: list[list[Any]],
    encoding: str = "utf-8",
) -> str:
    """Write dataset to CSV text.

    Parameters
    ----------
    headers : list[str]
        Column header labels.
    rows : list[list[Any]]
        Data rows (same order as *headers*).

    Returns
    -------
    str
        Full CSV as a single string.
    """
    buf = io.StringIO(newline="")
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue()


def to_xlsx(
    headers: list[str],
    rows: list[list[Any]],
    sheet_name: str = "Dataset",
    data_dict: list[dict[str, Any]] | None = None,
) -> bytes:
    """Write dataset to an Excel .xlsx file (in-memory ZIP/XLSX bytes).

    Parameters
    ----------
    headers, rows : dataset columns and data.
    sheet_name : str
        Name of the primary data sheet.
    data_dict : list[dict] or None
        Optional column-definition dicts for a "Data Dictionary" sheet.
        Each dict should have at least ``"variable"`` and ``"label"`` keys.

    Returns
    -------
    bytes
        Complete .xlsx file content (starts with ``PK`` ZIP signature).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row
    bold = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Optional data-dictionary sheet
    if data_dict:
        dd_ws = wb.create_sheet("Data Dictionary")
        dd_headers = list(data_dict[0].keys()) if data_dict else []
        for col_idx, h in enumerate(dd_headers, start=1):
            cell = dd_ws.cell(row=1, column=col_idx, value=h)
            cell.font = bold
        for row_idx, dd_row in enumerate(data_dict, start=2):
            for col_idx, key in enumerate(dd_headers, start=1):
                dd_ws.cell(row=row_idx, column=col_idx, value=dd_row.get(key, ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_sav(
    headers: list[str],
    rows: list[list[Any]],
    defs: dict[str, dict[str, Any]] | None = None,
) -> bytes:
    """Write dataset to an SPSS .sav file.

    Parameters
    ----------
    headers, rows : dataset columns and data.
    defs : dict
        Column definitions (from ``dictionary.column_defs()``).
        Keys are variable names; each value may contain:
        ``label`` (str), ``values`` (dict), ``measure`` (str), ``type`` (str).

    Returns
    -------
    bytes
        Complete .sav file content (starts with ``b"$FL2"``).
    """
    import pyreadstat

    # Build pyreadstat variable-value-label mapping
    variable_value_labels: dict[str, dict] = {}
    variable_measure: dict[str, str] = {}
    column_names: list[str] = []
    column_labels: list[str] = []

    for i, header in enumerate(headers):
        # Derive a clean variable name from the header
        var_name = f"var_{i}"
        if defs:
            # Try to find a matching key in defs (case-insensitive)
            found = False
            for key in defs:
                if key.lower() == header.lower().replace(" ", "_"):
                    var_name = key
                    found = True
                    break
            if not found and i < len(defs):
                # Fall back to positional matching
                var_name = list(defs.keys())[i]
        else:
            var_name = header.lower().replace(" ", "_")

        # Truncate to 64 chars (SPSS limit)
        var_name = var_name[:64]
        column_names.append(var_name)

        label = header
        if defs and var_name in defs:
            label = defs[var_name].get("label", header)
            val_map = defs[var_name].get("values", {})
            if val_map:
                variable_value_labels[var_name] = {
                    str(k): str(v) for k, v in val_map.items()
                }
            measure = defs[var_name].get("measure", "scale")
            variable_measure[var_name] = measure
        column_labels.append(label)

    # Build a pandas DataFrame with the resolved column names
    import pandas as pd

    df = pd.DataFrame(rows, columns=column_names)

    # Write to a temp file, then read back as bytes
    import tempfile, os

    tmp_path = None
    try:
        fd, tmp_path = tempfile.mkstemp(suffix=".sav")
        os.close(fd)
        pyreadstat.write_sav(
            df=df,
            dst_path=tmp_path,
            column_labels=column_labels or None,
            variable_value_labels=variable_value_labels or None,
            variable_measure=variable_measure or None,
        )
        with open(tmp_path, "rb") as f:
            sav_bytes = f.read()
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)

    return sav_bytes
